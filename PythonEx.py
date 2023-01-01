import csv
import math
import os
from pickletools import float8
import re
from decimal import Decimal, DecimalException
from operator import attrgetter
from statistics import mean
import math
from openpyxl import Workbook
from openpyxl.styles import Side, Font, Border, Alignment
from openpyxl.utils import get_column_letter

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }


class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = published_at
        self.salaryRUR = None
        self.year = 2010


class DataSet:

    def __init__(self, file_name):
        self.file_name = file_name
        self.dataVacancies = None
        self.vacancies_objects = None
        # (self.file_name)
        self.csv_to_dataset()

    def dataVacancies_to_vacancies_objects(self):
        listofvacansions = []
        for i in range(len(self.dataVacancies)):
            row = self.dataVacancies[i]

            vac = Vacancy(
                row['name']
                , row['salary_from']
                , row['salary_to']
                , row['salary_currency']
                , row['area_name']
                , row['published_at'])
            listofvacansions.append(vac)
        self.vacancies_objects = listofvacansions

    def csv_to_dataset(self):

        def csv_reader(file_name):
            with open(self.file_name, 'r', encoding="utf-8-sig") as file:
                reader = csv.reader(file)
                try:
                    headlines = next(reader)
                except StopIteration:
                    print("Пустой файл")
                return list(reader), headlines

        def csv_filer(reader, list_naming):
            updateInfo = {}
            newList = []
            for data in reader:
                if len(list_naming) == len(listNaming) and '' not in data:
                    for i in range(len(list_naming)):
                        if i == 2:
                            data[i] = re.sub(r'<[^<]+?>', '', data[i]).replace("\n", "; ").strip().split()
                        else:
                            data[i] = re.sub(r'<[^<]+?>', '', data[i]).replace("\n", ", ").strip().split()
                        data[i] = ' '.join(data[i])
                        updateInfo[list_naming[i]] = ''.join(data[i])
                    newList.append(updateInfo.copy())
                    updateInfo.clear()
            return newList

        reader2, listNaming = csv_reader(self.file_name)
        if len(reader2) != 0 and len(listNaming) != 0:
            self.dataVacancies = csv_filer(reader2, listNaming)
            self.dataVacancies_to_vacancies_objects()
        # dataset = DataSet(fileName,listofvacansions)
        # return dataset

        else:
            # print("Пустой файл1")
            # sys.exit()
            self.vacancies_objects = []


class CreateReport:
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    dataset = DataSet

    def __init__(self, dataset, professionName):
        self.dataset = dataset
        self.professionName = professionName
        self.vacbyyear = dict()
        self.countbyyear = dict()
        self.vacbyyearprof = dict()
        self.vacbyyearprofcount = dict()
        self.vacbytown = dict()
        self.citydictwithpart = dict()

        self.listOfYears = []
        self.city_dict = dict()
        # List с годами и расчет средней зарплаты из От,до и валюта
        self.fill_year_list_and_calc_salary1()
        # Динамика уровня зарплат по годам:
        self.vac_by_year()
        # Динамика количества вакансий по годам:
        self.vac_by_year_count()
        # Динамика уровня зарплат по годам для выбранной профессии:
        self.vac_by_year_prof()
        # Динамика количества вакансий по годам для выбранной профессии
        self.vac_by_year_prof_count()

        # Заполняем dict с городами
        self.fill_city_dict()
        # Уровень зарплат по городам (в порядке убывания
        self.salary_value_by_city()
        # Доля вакансий по городам (в порядке убывания)
        self.part_value_by_city()

    def fill_year_list_and_calc_salary(self):
        for vac in self.dataset.dataVacancies:
            vac['salaryRUR'] = self.mean_salary_in_rur(vac['salary_from'], vac['salary_to'], vac['salary_currency'])
            year = int(f'{vac["published_at"][0:4]}')
            vac['year'] = year
            self.listOfYears.append(year) if year not in self.listOfYears else self.listOfYears
            self.listOfYears.sort()

    def fill_year_list_and_calc_salary1(self):
        for vac in self.dataset.vacancies_objects:
            vac.salaryRUR = self.mean_salary_in_rur(vac.salary_from, vac.salary_to, vac.salary_currency)
            year = int(f'{vac.published_at[0:4]}')
            vac.year = year
            self.listOfYears.append(year) if year not in self.listOfYears else self.listOfYears
            self.listOfYears.sort()

    # расчет средней зарплаты в рублях

    def mean_salary_in_rur(self, salary_from, salary_to, salary_currency):
        currency_exchange_rate = float(self.currency_to_rub[salary_currency])
        return currency_exchange_rate * (float(salary_from) + float(salary_to)) / 2

    # Динамика уровня зарплат по годам:
    def vac_by_year(self):

        for y in self.listOfYears:
            # filtered = list(filter(lambda tag: int(tag['year']) == y , self.dataset.dataVacancies))
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            summa = sum(item.salaryRUR for item in filtered)
            records = len(filtered)
            means = int(summa / records)
            self.vacbyyear[y] = means
        print(f'Динамика уровня зарплат по годам: {self.vacbyyear}')

    # Динамика количества вакансий по годам:
    def vac_by_year_count(self):

        for y in self.listOfYears:
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            records = len(filtered)
            self.countbyyear[y] = records

        print(f'Динамика количества вакансий по годам: {self.countbyyear}')

    # Динамика уровня зарплат по годам для выбранной профессии:

    def vac_by_year_prof(self):

        for y in self.listOfYears:
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            filtered = list(filter(lambda tag: professionName in tag.name, filtered))

            summa = sum(item.salaryRUR for item in filtered)
            records = len(filtered)
            if len(filtered) > 0:
                means = int(summa / records)
                self.vacbyyearprof[y] = means
            else:
                self.vacbyyearprof[y] = 0

        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.vacbyyearprof}')

    # Динамика количества вакансий по годам для выбранной профессии
    def vac_by_year_prof_count(self):

        for y in self.listOfYears:
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            filtered = list(filter(lambda tag: professionName in tag.name, filtered))
            records = len(filtered)
            self.vacbyyearprofcount[y] = records

        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.vacbyyearprofcount}')

    # Добавляем из перечня вакансий key -Город value -количество вакансий
    def fill_city_dict(self):
        for vac in self.dataset.vacancies_objects:
            key = vac.area_name
            if key not in self.city_dict:
                self.city_dict[key] = 1
            else:
                self.city_dict[key] += 1
        # Удаляем те, где менее 1%
        records = len(self.dataset.vacancies_objects)
        for key, value in list(self.city_dict.items()):
            proc = math.floor(100 * value / records)
            if proc < 1:
                del self.city_dict[key]

    # Уровень зарплат по городам (в порядке убывания
    def salary_value_by_city(self):

        for city in self.city_dict:
            filtered = list(filter(lambda tag: tag.area_name == city, self.dataset.vacancies_objects))
            summa = sum(item.salaryRUR for item in filtered)
            records = len(filtered)
            means = int(summa / records)
            self.vacbytown[city] = means

        # Сортируем и берем 10 записей
        self.vacbytown = dict(sorted(self.vacbytown.items(), key=lambda item: item[1], reverse=True)[:10])
        print(f'Уровень зарплат по городам (в порядке убывания): {self.vacbytown}')

    # Доля вакансий по городам (в порядке убывания):
    def part_value_by_city(self):
        # Сортируем и берем 10 записей
        self.citydictwithpart = dict(sorted(self.city_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        records = len(self.dataset.vacancies_objects)
        # Расчитываем долю и пишем в value
        for key, value in list(self.citydictwithpart.items()):
            part = value / records
            self.citydictwithpart[key] = round(part, 4)

        print(f'Доля вакансий по городам (в порядке убывания): {self.citydictwithpart}')

    def get_info(self):
        return self.professionName, self.vacbyyear, self.countbyyear, self.vacbyyearprof, self.vacbyyearprofcount, self.vacbytown, self.citydictwithpart


class report():
    def __init__(self, professionName, vacbyyear, countbyyear, vacbyyearprof, vacbyyearprofcount, vacbytown,
                 citydictwithpart):
        self.professionName = professionName
        self.vacbyyear = vacbyyear
        self.countbyyear = countbyyear
        self.vacbyyearprof = vacbyyearprof
        self.vacbyyearprofcount = vacbyyearprofcount
        self.vacbytown = vacbytown
        self.citydictwithpart = citydictwithpart
        self.make_excel_report()

    def make_excel_report(self):
        workbook = Workbook()
        self.create_year_sheet(workbook)
        self.create_cities_sheet(workbook)
        self.sheet_formatting1(workbook)
        workbook.save('report.xlsx')

    def sheet_formatting1(self, workbook):
        side = Side(border_style='thin', color='000000')
        border = Border(left=side, top=side, right=side, bottom=side)
        for sheet in workbook.worksheets:
            # Жирный для первой строки
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)
            # ширина  колонок
            for index, column in enumerate(sheet.columns):
                if cell.row > 1 and cell.column == 5: cell.alignment = Alignment(horizontal='right')
                cell_width = 0
                for cell in column:

                    if cell.value:
                        cell.border = Border(left=side, top=side, right=side, bottom=side)
                        if len(str(cell.value)) + 1 > cell_width: cell_width = len(str(cell.value)) + 1
                    else:
                        cell_width = 2

                sheet.column_dimensions[get_column_letter(index + 1)].width = cell_width

            if sheet.title == 'Статистика по городам':
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
                    for cell in row:
                        cell.number_format = '0.00%'

    def create_year_sheet(self, workbook):
        headers = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.professionName}', 'Количество вакансий',
                   f'Количество вакансий - {self.professionName}']

        by_years = workbook.active
        by_years.title = 'Статистика по годам'
        by_years.append(headers)
        for index, key in enumerate(self.vacbyyear):
            by_years.cell(row=index + 2, column=1, value=key)
            by_years.cell(row=index + 2, column=2, value=self.vacbyyear[key])

        for index, key in enumerate(self.countbyyear):
            by_years.cell(row=index + 2, column=4, value=self.countbyyear[key])

        for index, key in enumerate(self.vacbyyearprof):
            by_years.cell(row=index + 2, column=3, value=self.vacbyyearprof[key])

        for index, key in enumerate(self.vacbyyearprofcount):
            by_years.cell(row=index + 2, column=5, value=self.vacbyyearprofcount[key])

    def create_cities_sheet(self, workbook):
        headers = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        by_cities = workbook.create_sheet('Статистика по городам')
        by_cities.title = 'Статистика по городам'
        by_cities.append(headers)
        for index, key in enumerate(self.vacbytown):
            by_cities.cell(row=index + 2, column=1, value=key)
            by_cities.cell(row=index + 2, column=2, value=self.vacbytown[key])

        for index, key in enumerate(self.citydictwithpart):
            by_cities.cell(row=index + 2, column=4, value=key)
            by_cities.cell(row=index + 2, column=5, value=self.citydictwithpart[key])


fileName = input("Введите название файла: ")
professionName = input("Введите название профессии: ")
dataset = DataSet(fileName)
cr = CreateReport(dataset, professionName)
professionName, vacbyyear, countbyyear, vacbyyearprof, vacbyyearprofcount, vacbytown, citydictwithpart = cr.get_info();
report(professionName, vacbyyear, countbyyear, vacbyyearprof, vacbyyearprofcount, vacbytown, citydictwithpart)

