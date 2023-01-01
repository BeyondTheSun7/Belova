import csv
import math
import os
from pickletools import float8
import re
from decimal import Decimal, DecimalException
from operator import attrgetter
from statistics import mean
import math
from openpyxl import Workbook
from openpyxl.styles import Side, Font, Border, Alignment
from openpyxl.utils import get_column_letter
from matplotlib import pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }


class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = published_at
        self.salaryRUR = None
        self.year = None


class DataSet:

    def __init__(self, file_name):
        self.file_name = file_name
        self.dataVacancies = None
        self.vacancies_objects = None
        # (self.file_name)
        self.csv_to_dataset()

    def dataVacancies_to_vacancies_objects(self):
        listofvacansions = []
        for i in range(len(self.dataVacancies)):
            row = self.dataVacancies[i]

            vac = Vacancy(
                row['name']
                , row['salary_from']
                , row['salary_to']
                , row['salary_currency']
                , row['area_name']
                , row['published_at'])
            listofvacansions.append(vac)
        self.vacancies_objects = listofvacansions

    def csv_to_dataset(self):

        def csv_reader(file_name):
            with open(self.file_name, 'r', encoding="utf-8-sig") as file:
                reader = csv.reader(file)
                try:
                    headlines = next(reader)
                except StopIteration:
                    print("Пустой файл")
                return list(reader), headlines

        def csv_filer(reader, list_naming):
            updateInfo = {}
            newList = []
            for data in reader:
                if len(list_naming) == len(listNaming) and '' not in data:
                    for i in range(len(list_naming)):
                        if i == 2:
                            data[i] = re.sub(r'<[^<]+?>', '', data[i]).replace("\n", "; ").strip().split()
                        else:
                            data[i] = re.sub(r'<[^<]+?>', '', data[i]).replace("\n", ", ").strip().split()
                        data[i] = ' '.join(data[i])
                        updateInfo[list_naming[i]] = ''.join(data[i])
                    newList.append(updateInfo.copy())
                    updateInfo.clear()
            return newList

        reader2, listNaming = csv_reader(self.file_name)
        if len(reader2) != 0 and len(listNaming) != 0:
            self.dataVacancies = csv_filer(reader2, listNaming)
            self.dataVacancies_to_vacancies_objects()
        else:
            self.vacancies_objects = []


class CreateData:
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    dataset = DataSet

    def __init__(self, dataset, professionName):
        self.dataset = dataset
        self.professionName = professionName
        self.vacbyyear = dict()
        self.countbyyear = dict()
        self.vacbyyearprof = dict()
        self.vacbyyearprofcount = dict()
        self.salary_level = dict()
        self.citydictwithpart = dict()

        self.listOfYears = []
        self.city_dict = dict()
        # List с годами и расчет средней зарплаты из От,до и валюта
        self.fill_year_list_and_calc_salary1()
        # Динамика уровня зарплат по годам:
        self.vac_by_year()
        # Динамика количества вакансий по годам:
        self.vac_by_year_count()
        # Динамика уровня зарплат по годам для выбранной профессии:
        self.vac_by_year_prof()
        # Динамика количества вакансий по годам для выбранной профессии
        self.vac_by_year_prof_count()

        # Заполняем dict с городами
        self.fill_city_dict()
        # Уровень зарплат по городам (в порядке убывания
        self.salary_value_by_city()
        # Доля вакансий по городам (в порядке убывания)
        self.part_value_by_city()

    def fill_year_list_and_calc_salary(self):
        for vac in self.dataset.dataVacancies:
            vac['salaryRUR'] = self.mean_salary_in_rur(vac['salary_from'], vac['salary_to'], vac['salary_currency'])
            year = int(f'{vac["published_at"][0:4]}')
            vac['year'] = year
            self.listOfYears.append(year) if year not in self.listOfYears else self.listOfYears
            self.listOfYears.sort()

    def fill_year_list_and_calc_salary1(self):
        for vac in self.dataset.vacancies_objects:
            vac.salaryRUR = self.mean_salary_in_rur(vac.salary_from, vac.salary_to, vac.salary_currency)
            year = int(f'{vac.published_at[0:4]}')
            vac.year = year
            self.listOfYears.append(year) if year not in self.listOfYears else self.listOfYears
            self.listOfYears.sort()

    # расчет средней зарплаты в рублях
    def mean_salary_in_rur(self, salary_from, salary_to, salary_currency):
        currency_exchange_rate = float(self.currency_to_rub[salary_currency])
        return currency_exchange_rate * (float(salary_from) + float(salary_to)) / 2

    # Динамика уровня зарплат по годам:
    def vac_by_year(self):

        for y in self.listOfYears:
            # filtered = list(filter(lambda tag: int(tag['year']) == y , self.dataset.dataVacancies))
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            summa = sum(item.salaryRUR for item in filtered)
            records = len(filtered)
            means = int(summa / records)
            self.vacbyyear[y] = means
        print(f'Динамика уровня зарплат по годам: {self.vacbyyear}')

    # Динамика количества вакансий по годам:
    def vac_by_year_count(self):
        for y in self.listOfYears:
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            records = len(filtered)
            self.countbyyear[y] = records
        print(f'Динамика количества вакансий по годам: {self.countbyyear}')

    # Динамика уровня зарплат по годам для выбранной профессии:
    def vac_by_year_prof(self):
        for y in self.listOfYears:
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            filtered = list(filter(lambda tag: professionName in tag.name, filtered))

            summa = sum(item.salaryRUR for item in filtered)
            records = len(filtered)
            if len(filtered) > 0:
                means = int(summa / records)
                self.vacbyyearprof[y] = means
            else:
                self.vacbyyearprof[y] = 0

        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.vacbyyearprof}')

    # Динамика количества вакансий по годам для выбранной профессии
    def vac_by_year_prof_count(self):
        for y in self.listOfYears:
            filtered = list(filter(lambda tag: tag.year == y, self.dataset.vacancies_objects))
            filtered = list(filter(lambda tag: professionName in tag.name, filtered))
            records = len(filtered)
            self.vacbyyearprofcount[y] = records
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.vacbyyearprofcount}')

    # Добавляем из перечня вакансий key -Город value -количество вакансий
    def fill_city_dict(self):
        for vac in self.dataset.vacancies_objects:
            key = vac.area_name
            if key not in self.city_dict:
                self.city_dict[key] = 1
            else:
                self.city_dict[key] += 1
        # Удаляем те, где менее 1%
        records = len(self.dataset.vacancies_objects)
        for key, value in list(self.city_dict.items()):
            proc = math.floor(100 * value / records)
            if proc < 1:
                del self.city_dict[key]

    # Уровень зарплат по городам (в порядке убывания
    def salary_value_by_city(self):
        for city in self.city_dict:
            filtered = list(filter(lambda tag: tag.area_name == city, self.dataset.vacancies_objects))
            summa = sum(item.salaryRUR for item in filtered)
            records = len(filtered)
            means = int(summa / records)
            self.salary_level[city] = means

        # Сортируем и берем 10 записей
        self.salary_level = dict(sorted(self.salary_level.items(), key=lambda item: item[1], reverse=True)[:10])
        print(f'Уровень зарплат по городам (в порядке убывания): {self.salary_level}')

    # Доля вакансий по городам (в порядке убывания):
    def part_value_by_city(self):
        # Сортируем и берем 10 записей
        self.citydictwithpart = dict(sorted(self.city_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        records = len(self.dataset.vacancies_objects)
        # Расчитываем долю и пишем в value
        for key, value in list(self.citydictwithpart.items()):
            part = value / records
            self.citydictwithpart[key] = round(part, 4)

        print(f'Доля вакансий по городам (в порядке убывания): {self.citydictwithpart}')

    def get_info(self):
        return self.professionName, self.vacbyyear, self.countbyyear, self.vacbyyearprof, self.vacbyyearprofcount, self.salary_level, self.citydictwithpart


class report():
    def __init__(self, professionName, vacbyyear, countbyyear, vacbyyearprof, vacbyyearprofcount, salary_level,
                 citydictwithpart):
        self.professionName = professionName
        self.vacbyyear = vacbyyear
        self.countbyyear = countbyyear
        self.vacbyyearprof = vacbyyearprof
        self.vacbyyearprofcount = vacbyyearprofcount
        self.salary_level = salary_level
        self.citydictwithpart = citydictwithpart

    def generate_excel(self):
        workbook = Workbook()
        self.create_year_sheet(workbook)
        self.create_cities_sheet(workbook)
        self.sheet_formatting1(workbook)
        workbook.save('report.xlsx')

    def sheet_formatting1(self, workbook):
        side = Side(border_style='thin', color='000000')
        border = Border(left=side, top=side, right=side, bottom=side)
        for sheet in workbook.worksheets:
            # Жирный для первой строки
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)
            # ширина  колонок
            for index, column in enumerate(sheet.columns):
                if cell.row > 1 and cell.column == 5: cell.alignment = Alignment(horizontal='right')
                cell_width = 0
                for cell in column:

                    if cell.value:
                        cell.border = Border(left=side, top=side, right=side, bottom=side)
                        if len(str(cell.value)) + 1 > cell_width: cell_width = len(str(cell.value)) + 1
                    else:
                        cell_width = 2

                sheet.column_dimensions[get_column_letter(index + 1)].width = cell_width

            if sheet.title == 'Статистика по городам':
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
                    for cell in row:
                        cell.number_format = '0.00%'

    def create_year_sheet(self, workbook):
        headers = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.professionName}', 'Количество вакансий',
                   f'Количество вакансий - {self.professionName}']

        by_years = workbook.active
        by_years.title = 'Статистика по годам'
        by_years.append(headers)
        for index, key in enumerate(self.vacbyyear):
            by_years.cell(row=index + 2, column=1, value=key)
            by_years.cell(row=index + 2, column=2, value=self.vacbyyear[key])

        for index, key in enumerate(self.countbyyear):
            by_years.cell(row=index + 2, column=4, value=self.countbyyear[key])

        for index, key in enumerate(self.vacbyyearprof):
            by_years.cell(row=index + 2, column=3, value=self.vacbyyearprof[key])

        for index, key in enumerate(self.vacbyyearprofcount):
            by_years.cell(row=index + 2, column=5, value=self.vacbyyearprofcount[key])

    def create_cities_sheet(self, workbook):
        headers = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        by_cities = workbook.create_sheet('Статистика по городам')
        by_cities.title = 'Статистика по городам'
        by_cities.append(headers)
        for index, key in enumerate(self.salary_level):
            by_cities.cell(row=index + 2, column=1, value=key)
            by_cities.cell(row=index + 2, column=2, value=self.salary_level[key])

        for index, key in enumerate(self.citydictwithpart):
            by_cities.cell(row=index + 2, column=4, value=key)
            by_cities.cell(row=index + 2, column=5, value=self.citydictwithpart[key])

    def generate_image(self):
        fig, axs = plt.subplots(2, 2, figsize=(9, 6), constrained_layout=True)

        labels = list(self.vacbyyear.keys())
        year_means = list(self.vacbyyear.values())
        proff_means = list(self.vacbyyearprof.values())

        x = np.arange(len(labels))  # the label locations
        width = 0.35  # the width of the bars
        ax1 = axs[0, 0]
        ax1.bar(x - width / 2, year_means, width, label='Средняя з/п')
        ax1.bar(x + width / 2, proff_means, width, label=f'з/п {professionName}')
        ax1.set_title('Уровень зарплат по годам')
        ax1.tick_params(axis='x', labelrotation=90)
        ax1.set_xticks(x, labels)
        ax1.grid(axis='y')
        ax1.legend(loc='best', fontsize=8)
        for item in ([ax1.xaxis.label, ax1.yaxis.label] +
                     ax1.get_xticklabels() + ax1.get_yticklabels()):
            item.set_fontsize(8)

        # количество вакансий по годам как общий, так и для выбранной профессии
        labels1 = list(self.countbyyear.keys())
        year_means = list(self.countbyyear.values())
        proff_means = list(self.vacbyyearprofcount.values())

        x = np.arange(len(labels1))  # the label locations
        width = 0.35  # the width of the bars
        ax2 = axs[0, 1]
        ax2.bar(x - width / 2, year_means, width, label='Количество вакансий')
        ax2.bar(x + width / 2, proff_means, width, label=f'Количество вакансий {professionName}')
        ax2.set_title('Количество вакансий по годам')
        ax2.tick_params(axis='x', labelrotation=90)
        ax2.set_xticks(x, labels1)
        ax2.grid(axis='y')
        ax2.legend(loc='best', fontsize=8)
        for item in ([ax2.xaxis.label, ax2.yaxis.label] +
                     ax2.get_xticklabels() + ax2.get_yticklabels()):
            item.set_fontsize(8)

        # горизонтальная диаграмма - уровень зарплат по городам:
        labels2 = list(self.salary_level.keys())
        # Иначе неверный порядок
        labels2.reverse()
        labels2 = [sub.replace(' ', '\n') for sub in labels2]
        labels2 = [sub.replace('-', '-\n') for sub in labels2]

        salary_level = list(self.salary_level.values())
        salary_level.reverse()

        ax3 = axs[1, 0]
        ax3.barh(labels2, salary_level)
        ax3.set_title('Уровень зарплат по городам')

        for item in ax3.get_yticklabels():
            item.set_fontsize(6)
            item.set_verticalalignment('center')
            item.set_horizontalalignment('right')

        for item in ([ax3.xaxis.label, ax3.yaxis.label] +
                     ax3.get_xticklabels()):
            item.set_fontsize(8)

        # круговая диаграмма - доля вакансий по городам.
        otherpart = 1 - sum(self.citydictwithpart.values())
        self.citydictwithpart["Другие"] = otherpart
        self.citydictwithpart = dict(sorted(self.citydictwithpart.items(), key=lambda item: item[1], reverse=True))
        labels3 = list(self.citydictwithpart.keys())
        sizes = list(self.citydictwithpart.values())
        ax4 = axs[1, 1]
        ax4.set_title('Доля вакансий по городам')
        ax4.pie(sizes, labels=labels3,
                shadow=False, startangle=0, textprops={'fontsize': 6})
        ax4.axis('equal')
        plt.savefig("fig1.png", dpi=300)

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("template.html")
        statby_years = []
        salary_level_data = []
        part_by_city = []
        self.listOfYears = []

        for key in self.vacbyyear:
            statby_years.append(
                [key, self.vacbyyear[key], self.vacbyyearprof[key], self.countbyyear[key], vacbyyearprofcount[key]])
        for key in self.salary_level:
            salary_level_data.append([key, self.salary_level[key]])
        for key in self.citydictwithpart:
            if key != "Другие":
                part_by_city.append([key, f'{round(self.citydictwithpart[key] * 100, 2)}%'])

        table1columns = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.professionName}',
                         'Количество вакансий', f'Количество вакансий - {self.professionName}']
        table2columns = ["Город", "Уровень зарплат"]
        table3columns = ["Город", "Доля вакансий"]
        pdf_template = template.render(
            {'table1columns': table1columns, 'statby_years': statby_years, 'name': self.professionName,
             'table2columns': table2columns, 'table3columns': table3columns,
             'salary_level': salary_level_data, 'part_by_city': part_by_city})
        config = pdfkit.configuration(wkhtmltopdf=r'd:\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'out.pdf', configuration=config, options={'enable-local-file-access': ''})
        return


fileName = input("Введите название файла: ")
professionName = input("Введите название профессии: ")
fileName = "vacancies_by_year.csv"

dataset = DataSet(fileName)
data = CreateData(dataset, professionName)
professionName, vacbyyear, countbyyear, vacbyyearprof, vacbyyearprofcount, salary_level, citydictwithpart = data.get_info();
re = report(professionName, vacbyyear, countbyyear, vacbyyearprof, vacbyyearprofcount, salary_level, citydictwithpart)
re.generate_excel()
re.generate_image()
re.generate_pdf()